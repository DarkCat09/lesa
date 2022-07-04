import os
import csv
import enum
import openpyxl
from zipfile import ZipFile
from datetime import date, timedelta
from tempfile import NamedTemporaryFile
from flask import redirect
from flask import make_response
from typing import Any

from ..app import app
from ..upload import photos_dir
from ..consts import XLSX_MIME
from ..models import Post, Register
from ..models import Parent, Child

from . import check_token

class CsvType(enum.IntEnum):
    main = 0
    parents = 1
    children = 2

@app.route('/admin/backup/reg/csv')
def reg_csv_main():
    return reg_csv(CsvType.main)

@app.route('/admin/backup/regp/csv')
def reg_csv_parents():
    return reg_csv(CsvType.parents)

@app.route('/admin/backup/regc/csv')
def reg_csv_children():
    return reg_csv(CsvType.children)

@app.route('/admin/backup/reg/xls')
def regdb_xlsx_main():
    return regdb_xlsx(add_main=True)

@app.route('/admin/backup/regp/xls')
def regdb_xlsx_parents():
    return regdb_xlsx(add_parents=True)

@app.route('/admin/backup/regc/xls')
def regdb_xlsx_children():
    return regdb_xlsx(add_children=True)

@app.route('/admin/backup/regdb/xls')
def regdb_xlsx_full():
    return regdb_xlsx(
        add_main=True,
        add_parents=True,
        add_children=True
    )

def reg_csv(csv_type:CsvType) -> Any:

    if not check_token():
        return redirect('/admin')

    resp = ''
    with NamedTemporaryFile('wt+', newline='') as file:
        table = csv.writer(file, delimiter=';')

        if csv_type == CsvType.main:

            regdb = Register.query.all()

            table.writerow((
                '№', 'ID', 'Семья', 'Смена',
                'Кол-во', 'Взрослые', 'Дети',
                'Дом', 'Друзья'
            ))

            for r in regdb:
                adults = int(r.count or 0) - int(r.children or 0)
                table.writerow((
                    r.id, r.mid, r.family, r.dates,
                    r.count, adults, r.children,
                    r.house, r.friends
                ))
        
        if csv_type == CsvType.parents:

            parents = Parent.query.all()

            table.writerow((
                '№', 'ID', 'Фамилия', 'Имя', 'Отчество',
                'Телефон', 'Почта', 'Соц. сеть'
            ))

            for p in parents:
                table.writerow((
                    p.id, p.mid, p.surname, p.firstname, p.midname,
                    p.phone, p.email, p.social
                ))
        
        if csv_type == CsvType.children:

            children = Child.query.all()

            table.writerow((
                '№', 'ID', 'Фамилия', 'Имя',
                'Пол', 'Дата рождения', 'Возраст'
            ))

            for c in children:
                age:timedelta = date.today() - c.birthday
                table.writerow((
                    c.id, c.mid, c.surname, c.firstname,
                    c.gender, c.birthday, age.days // 365
                ))
                del age
        
        file.seek(0)
        resp = make_response(file.read())
        resp.content_type = 'text/csv'
        resp.content_length = file.tell()
    
    return resp

def regdb_xlsx(
    add_main:bool=False,
    add_parents:bool=False,
    add_children:bool=False) -> Any:

    if not check_token():
        return redirect('/admin')
    
    regdb = Register.query.all()
    parents = Parent.query.all()
    children = Child.query.all()

    table = openpyxl.Workbook()
    table.remove(table.active)

    ws_main = []
    if add_main:
        ws_main = table.create_sheet('Основные данные')

    ws_parents = []
    if add_parents:
        ws_parents = table.create_sheet('Родители')

    ws_children = []
    if add_children:
        ws_children = table.create_sheet('Дети')

    if add_main:
        ws_main.append((
            '№', 'ID', 'Семья', 'Смена',
            'Кол-во', 'Взрослые', 'Дети',
            'Дом', 'Друзья'
        ))

    if add_parents:
        ws_parents.append((
            '№', 'ID', 'Фамилия', 'Имя', 'Отчество',
            'Телефон', 'Почта', 'Соц. сеть'
        ))
    
    if add_children:
        ws_children.append((
            '№', 'ID', 'Фамилия', 'Имя',
            'Пол', 'Дата рождения', 'Возраст'
        ))

    for r in regdb:

        if add_main:
            adults = int(r.count or 0) - int(r.children or 0)
            ws_main.append((
                r.id, r.mid, r.family, r.dates,
                r.count, adults, r.children,
                r.house, r.friends
            ))

        if add_parents:
            f_parents = [p for p in parents if p.mid == r.mid]
            for p in f_parents:
                ws_parents.append((
                    p.id, p.mid, p.surname, p.firstname, p.midname,
                    p.phone, p.email, p.social
                ))
        
        if add_children:
            f_children = [c for c in children if c.mid == r.mid]
            for c in f_children:
                age:timedelta = date.today() - c.birthday
                ws_children.append((
                    c.id, c.mid, c.surname, c.firstname,
                    c.gender, c.birthday, age.days // 365
                ))
                del age

    resp = ''
    with NamedTemporaryFile('wb+') as file:
        table.save(file.name)
        file.seek(0)
        resp = make_response(file.read())
        resp.content_type = XLSX_MIME
        resp.content_length = file.tell()
    return resp

@app.route('/admin/backup/posts/csv')
def posts_csv():
    
    if not check_token():
        return redirect('/admin')

    resp = ''
    with NamedTemporaryFile('wt+', newline='') as file:
        table = csv.writer(file, delimiter=';')
        posts = Post.query.all()
        table.writerow((
            '№', 'Заголовок',
            'Unix-время', 'Время',
            'Содержимое'
        ))

        for p in posts:
            timestamp = p.dt.timestamp()
            table.writerow((
                p.id, p.title,
                timestamp, p.dt,
                p.body
            ))
        
        file.seek(0)
        resp = make_response(file.read())
        resp.content_type = 'text/csv'
        resp.content_length = file.tell()
    
    return resp

@app.route('/admin/backup/images/zip')
def images_zip():

    resp = ''
    with NamedTemporaryFile('wb+') as file:

        with ZipFile(file, 'w') as zip:
            for root, dirs, files in os.walk(photos_dir):
                for path in files:

                    # relative to photos_dir
                    reldir = root.replace(photos_dir, '')
                    relpath = os.path.join(reldir, path)
                    # full
                    fullpath = os.path.join(root, path)

                    zip.write(fullpath, relpath)

        file.seek(0)
        resp = make_response(file.read())
        resp.content_type = 'application/zip'
        resp.content_length = file.tell()

    return resp
